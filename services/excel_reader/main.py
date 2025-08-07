from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook
import json
from io import BytesIO

app = FastAPI(title="Excel Reader", version="0.1.0")

@app.get("/health")
def health():
    return {"status": "healthy", "service": "excel_reader"}

@app.post("/snapshot")
async def snapshot(file: UploadFile = File(...)):
    content = await file.read()
    
    # Загружаем книгу дважды - для значений и для формул
    wb_values = load_workbook(BytesIO(content), data_only=True)
    wb_formulas = load_workbook(BytesIO(content), data_only=False)
    
    def generate():
        for sheet_idx, sheet_values in enumerate(wb_values.worksheets):
            sheet_formulas = wb_formulas.worksheets[sheet_idx]
            
            for row_idx, row in enumerate(sheet_values.iter_rows()):
                for col_idx, cell in enumerate(row):
                    # Получаем соответствующую ячейку с формулой
                    cell_formula = sheet_formulas.cell(row=row_idx+1, column=col_idx+1)
                    
                    # Если есть значение или формула
                    if cell.value is not None or (hasattr(cell_formula, 'value') and 
                                                 isinstance(cell_formula.value, str) and 
                                                 cell_formula.value.startswith('=')):
                        
                        formula = None
                        if hasattr(cell_formula, 'value') and isinstance(cell_formula.value, str) and cell_formula.value.startswith('='):
                            formula = cell_formula.value
                        
                        yield json.dumps({
                            "sheet": sheet_values.title,
                            "addr": f"{sheet_values.title}!{cell.coordinate}",
                            "value": cell.value,
                            "formula": formula
                        }) + "\n"
    
    return StreamingResponse(generate(), media_type="application/x-ndjson")