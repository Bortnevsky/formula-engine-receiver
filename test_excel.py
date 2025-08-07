import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test"
ws["A1"] = "Revenue"
ws["B1"] = 1000
ws["A2"] = "Cost"
ws["B2"] = 600
ws["A3"] = "Profit"
ws["B3"] = "=B1-B2"

wb.save("data/test.xlsx")
print("Test Excel created!")